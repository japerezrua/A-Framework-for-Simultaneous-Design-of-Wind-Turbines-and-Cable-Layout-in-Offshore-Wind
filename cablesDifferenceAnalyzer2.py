# -*- coding: utf-8 -*-
"""
Created on Thu Mar 11 13:55:15 2021

@author: juru
"""

import openpyxl
import numpy as np
import statistics
import matplotlib.pylab as plt
import plotly.express as px
import pandas as pd
#%% Inputs
file_name='17-05-2021_13-05-14_Processing_.xlsx'
diameter=190.6 #Diameter in meters

#%% Process
wb1 = openpyxl.load_workbook(file_name,data_only=True)
sheet1 = wb1['Outputs']
sheet2 = wb1['Cables outputs']
Time_array,Length_av_array,Length_sd_array=np.empty(0),np.empty(0),np.empty(0)
Perc_array,Heu_array,Glob_array=np.empty(0),np.empty(0),np.empty(0)
of_d_array,of_g_array,ide_array,aep_array=np.empty(0),np.empty(0),np.empty(0),np.empty(0)
file_name_1=[]

for i in range(2,5000,1):
    File=sheet1.cell(row=i, column=1).value
    if File is None:
       break
    Time=sheet1.cell(row=i, column=2).value
    Length_av=sheet1.cell(row=i, column=3).value
    Length_sd=sheet1.cell(row=i, column=4).value
    Perc=sheet1.cell(row=i, column=5).value
    Heu=sheet1.cell(row=i, column=6).value
    Glob=sheet1.cell(row=i, column=7).value
    aep=sheet1.cell(row=i, column=8).value
    of_d=sheet1.cell(row=i, column=9).value
    of_g=sheet1.cell(row=i, column=10).value
    ide=sheet1.cell(row=i, column=11).value
    Time_array=np.concatenate((Time_array,np.array(Time).reshape(1)))
    Length_av_array=np.concatenate((Length_av_array,np.array(Length_av).reshape(1)))
    Length_sd_array=np.concatenate((Length_sd_array,np.array(Length_sd).reshape(1)))
    Perc_array=np.concatenate((Perc_array,np.array(Perc).reshape(1)))
    Heu_array=np.concatenate((Heu_array,np.array(Heu).reshape(1)))
    Glob_array=np.concatenate((Glob_array,np.array(Glob).reshape(1)))
    aep_array=np.concatenate((aep_array,np.array(aep).reshape(1)))
    of_d_array=np.concatenate((of_d_array,np.array(of_d).reshape(1)))
    of_g_array=np.concatenate((of_g_array,np.array(of_g).reshape(1)))
    ide_array=np.concatenate((ide_array,np.array(ide).reshape(1)))
    file_name_1=file_name_1+[File]
ide_array_2, global_500, global_800, global_1000=np.empty(0),np.empty(0),np.empty(0),np.empty(0)
heuris_500, heuris_800, heuris_1000=np.empty(0),np.empty(0),np.empty(0)
file_name_2=[]
Length_av_array=Length_av_array/diameter
Length_sd_array=Length_sd_array/diameter

for i in range(2,5000,1):
    File=sheet2.cell(row=i, column=1).value
    if File is None:
       break
    ide_array_2_val=sheet2.cell(row=i, column=2).value
    global_500_val=sheet2.cell(row=i, column=3).value
    global_800_val=sheet2.cell(row=i, column=4).value
    global_1000_val=sheet2.cell(row=i, column=5).value
    heuris_500_val=sheet2.cell(row=i, column=6).value
    heuris_800_val=sheet2.cell(row=i, column=7).value
    heuris_1000_val=sheet2.cell(row=i, column=8).value
    ide_array_2=np.concatenate((ide_array_2,np.array(ide_array_2_val).reshape(1)))
    global_500=np.concatenate((global_500,np.array(global_500_val).reshape(1)))
    global_800=np.concatenate((global_800,np.array(global_800_val).reshape(1)))
    global_1000=np.concatenate((global_1000,np.array(global_1000_val).reshape(1)))
    heuris_500=np.concatenate((heuris_500,np.array(heuris_500_val).reshape(1)))
    heuris_800=np.concatenate((heuris_800,np.array(heuris_800_val).reshape(1)))
    heuris_1000=np.concatenate((heuris_1000,np.array(heuris_1000_val).reshape(1)))
    file_name_2=file_name_2+[File]
#%% Output displaying
fig = plt.figure(figsize=(9, 10))
ax = fig.add_subplot(111, projection='3d')
ini0,ini1,ini2,ini3=-1,-1,-1,-1
for i in range(len(Time_array)):
    if (ide_array[i]==0) and (ini0==0):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='b', marker='o')
    if (ide_array[i]==0) and (ini0==-1):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='b', marker='o',label='OSS at approximately centroid, Layouts from Approach 1')
       ini0=0
    if (ide_array[i]==1) and (ini1==0):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='r', marker='o')
    if (ide_array[i]==1) and (ini1==-1):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='r', marker='o',label='OSS at approximately centroid, Layouts from Approach 2')
       ini1=0    
    if (ide_array[i]==2) and (ini2==0):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='b', marker='^')
    if (ide_array[i]==2) and (ini2==-1):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='b', marker='^',label='OSS at external location, Layouts from Approach 1')
       ini2=0
    if (ide_array[i]==3) and (ini3==0):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='r', marker='^')
    if (ide_array[i]==3) and (ini3==-1):
       ax.scatter(Length_av_array[i],Length_sd_array[i],Perc_array[i],c='r', marker='^',label='OSS at external location, Layouts from Approach 2')
       ini3=0            
ax.set_xlabel('Average length between pairs of WTs [D]')
ax.set_ylabel('Standard dev. length between pairs of WTs [D]')
ax.set_zlabel('Abs. relative cost diff. between heuristic and global optimizer [%]')
#Add legends!
plt.legend(loc='upper right')
plt.show()    
#%% Analyzing
"""
print("Pearson coefficient between average length and SD length:",np.corrcoef(Length_av_array,Length_sd_array)[0,1])
plt.figure()
plt.plot(Length_av_array,Length_sd_array,'o')
m, b = np.polyfit(Length_av_array,Length_sd_array, 1)
plt.plot(Length_av_array, m*Length_av_array + b)
plt.xlabel('Average length [m]')
plt.ylabel('SD length [m]')
plt.title('Correlation average length and SD length')
plt.text(7250,4600,'r='+str(round(np.corrcoef(Length_av_array,Length_sd_array)[0,1],2)))

percentile=50
percentile_c=np.percentile(Perc_array,percentile)
print("Percentile",percentile,"th of absolute relative difference between heuristic and global optimzer:",percentile_c)
#i_near=abs(Perc_array-percentile_c).argmin()
plt.figure()
plt.plot(Glob_array[Perc_array<=percentile_c],Heu_array[Perc_array<=percentile_c],'o')
m, b = np.polyfit(Glob_array[Perc_array<=percentile_c],Heu_array[Perc_array<=percentile_c], 1)
plt.plot(Glob_array[Perc_array<=percentile_c], m*Glob_array[Perc_array<=percentile_c] + b)
plt.xlabel('Cost from global optimizer [Euros]')
plt.ylabel('Cost from heuristic [Euros]')
#plt.title('Correlation global optimizer and heuristic for 50th percentile of cost difference')
plt.text(0.85e8,1.3e8,'r='+str(round(np.corrcoef(Glob_array[Perc_array<=percentile_c],Heu_array[Perc_array<=percentile_c])[0,1],2)))
print("Pearson coefficient between heuristic and global optimizer below percentile of difference:",np.corrcoef(Glob_array[Perc_array<=percentile_c],Heu_array[Perc_array<=percentile_c])[0,1])

plt.figure()
plt.plot(Glob_array[Perc_array>percentile_c],Heu_array[Perc_array>percentile_c],'o')
m, b = np.polyfit(Glob_array[Perc_array>percentile_c],Heu_array[Perc_array>percentile_c], 1)
plt.plot(Glob_array[Perc_array>percentile_c], m*Glob_array[Perc_array>percentile_c] + b)
plt.xlabel('Cost from global optimizer [Euros]')
plt.ylabel('Cost from heuristic [Euros]')
plt.title('Correlation global optimizer and heuristic for 50th to 100th percentile of cost difference')
plt.text(1e8,2e8,'r='+str(round(np.corrcoef(Glob_array[Perc_array>percentile_c],Heu_array[Perc_array>percentile_c])[0,1],2)))
print("Pearson coefficient between heuristic and global optimizer above percentile of difference:",np.corrcoef(Glob_array[Perc_array>percentile_c],Heu_array[Perc_array>percentile_c])[0,1])

percentile=50
percentile_l=np.percentile(Length_av_array,percentile)
print("Percentile",percentile,"th of average length:",percentile_l)

plt.figure()
plt.plot(Glob_array[Length_av_array<=percentile_l],Heu_array[Length_av_array<=percentile_l],'o')
m, b = np.polyfit(Glob_array[Length_av_array<=percentile_l],Heu_array[Length_av_array<=percentile_l], 1)
plt.plot(Glob_array[Length_av_array<=percentile_l], m*Glob_array[Length_av_array<=percentile_l] + b)
plt.xlabel('Cost from global optimizer [Euros]')
plt.ylabel('Cost from heuristic [Euros]')
plt.title('Correlation global optimizer and heuristic for 50th percentile of average length')
plt.text(0.85e8,1.6e8,'r='+str(round(np.corrcoef(Glob_array[Length_av_array<=percentile_l],Heu_array[Length_av_array<=percentile_l])[0,1],2)))
print("Pearson coefficient between heuristic and global optimizer below percentile of length:",np.corrcoef(Glob_array[Length_av_array<=percentile_l],Heu_array[Length_av_array<=percentile_l])[0,1])
#Glob_array[np.argwhere(np.logical_and(ide_array==2,Heu_array<100000000))].shape

plt.figure()
plt.plot(Glob_array[Length_av_array>percentile_l],Heu_array[Length_av_array>percentile_l],'o')
m, b = np.polyfit(Glob_array[Length_av_array>percentile_l],Heu_array[Length_av_array>percentile_l], 1)
plt.plot(Glob_array[Length_av_array>percentile_l], m*Glob_array[Length_av_array>percentile_l] + b)
plt.xlabel('Cost from global optimizer [Euros]')
plt.ylabel('Cost from heuristic [Euros]')
plt.title('Correlation global optimizer and heuristic for 50th to 100th percentile of average length')
plt.text(1.05e8,2.1e8,'r='+str(round(np.corrcoef(Glob_array[Length_av_array>percentile_l],Heu_array[Length_av_array>percentile_l])[0,1],2)))
print("Pearson coefficient between heuristic and global optimizer above percentile of length:",np.corrcoef(Glob_array[Length_av_array>percentile_l],Heu_array[Length_av_array>percentile_l])[0,1])
"""

np.array(file_name_2)[ide_array_2==0]
global_500[ide_array_2==0]
global_800[ide_array_2==0]
global_1000[ide_array_2==0]
heuris_500[ide_array_2==0]
heuris_800[ide_array_2==0]
heuris_1000[ide_array_2==0]
dev_500_A1=(heuris_500[ide_array_2==0]-global_500[ide_array_2==0])*100/(global_500[ide_array_2==0])
dev_500_A2=(heuris_500[ide_array_2==1]-global_500[ide_array_2==1])*100/(global_500[ide_array_2==1])

dev_800_A1=(heuris_800[ide_array_2==0]-global_800[ide_array_2==0])*100/(global_800[ide_array_2==0])
dev_800_A2=(heuris_800[ide_array_2==1]-global_800[ide_array_2==1])*100/(global_800[ide_array_2==1])

dev_1000_A1=(heuris_1000[ide_array_2==0]-global_1000[ide_array_2==0])*100/(global_1000[ide_array_2==0])
dev_1000_A2=(heuris_1000[ide_array_2==1]-global_1000[ide_array_2==1])*100/(global_1000[ide_array_2==1])

total_length_H_A1=heuris_500[ide_array_2==0]+heuris_800[ide_array_2==0]+heuris_1000[ide_array_2==0]
total_length_G_A1=global_500[ide_array_2==0]+global_800[ide_array_2==0]+global_1000[ide_array_2==0]

total_length_H_A2=heuris_500[ide_array_2==1]+heuris_800[ide_array_2==1]+heuris_1000[ide_array_2==1]
total_length_G_A2=global_500[ide_array_2==1]+global_800[ide_array_2==1]+global_1000[ide_array_2==1]

print("Average difference on length of cable 500mm2 between heuristic and global w.r.t global in Approach 1:",np.average(dev_500_A1))
print("Average difference on length of cable 800mm2 between heuristic and global w.r.t global in Approach 1:",np.average(dev_800_A1))
print("Average difference on length of cable 1000mm2 between heuristic and global w.r.t global in Approach 1:",np.average(dev_1000_A1))
print("Average total length of cables with heuristic in Approach 1:",np.average(total_length_H_A1))
print("Average total length of cables with global in Approach 1:",np.average(total_length_G_A1))

print("Average difference on length of cable 500mm2 between heuristic and global w.r.t global in Approach 2:",np.average(dev_500_A2))
print("Average difference on length of cable 800mm2 between heuristic and global w.r.t global in Approach 2:",np.average(dev_800_A2))
print("Average difference on length of cable 1000mm2 between heuristic and global w.r.t global in Approach 2:",np.average(dev_1000_A2))
print("Average total length of cables with heuristic in Approach 2:",np.average(total_length_H_A2))
print("Average total length of cables with global in Approach 2:",np.average(total_length_G_A2))

labels = ['Cable type 1 (4 WTs)', 'Cable type 2 (6 WTs)', 'Cable type 3 (8 WTs)']
Approach1=[round(np.average(dev_500_A1),2),round(np.average(dev_800_A1),2),round(np.average(dev_1000_A1),2)]
Approach2=[round(np.average(dev_500_A2),2),round(np.average(dev_800_A2),2),round(np.average(dev_1000_A2),2)]
x = np.arange(len(labels))
width = 0.20
fig, ax = plt.subplots()
rects1 = ax.bar(x - width/2, Approach1, width, label='Layouts from Approach 1',color='b')
rects2 = ax.bar(x + width/2, Approach2, width, label='Layouts from Approach 2',color='r')
ax.set_ylabel('Average relative length diff. between heuristic and global optimizer [%]',fontsize=18)
ax.set_xticks(x)
ax.set_xticklabels(labels,fontsize=18)
ax.tick_params(axis="y", labelsize=18)
ax.legend(fontsize=18)
def autolabel(rects):
    for rect in rects:
        height = rect.get_height()
        if height>=0:
            ax.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom',fontsize=18)
        if height<0:
            ax.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, -22),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom',fontsize=18)           
autolabel(rects1)
autolabel(rects2) 
fig.tight_layout()
ax.axhline(y=0,color="black",linewidth=0.5)
ax.annotate("Total length difference [%]: {total}".format(total = round(100*(np.average(total_length_H_A1)-np.average(total_length_G_A1))/(np.average(total_length_G_A1)),2)),xy=(-0.3,325),color='b',fontsize=18)
ax.annotate("Total length difference [%]: {total}".format(total = round(100*(np.average(total_length_H_A2)-np.average(total_length_G_A2))/(np.average(total_length_G_A2)),2)),xy=(-0.3,310),color='r',fontsize=18)
plt.show()

labels = ['Approach 1','Approach 2']
x = np.arange(len(labels))
Approach1_o=of_g_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==0)))]
Approach2_o=of_g_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==1)))]
Approach1_e=of_g_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==2)))]
Approach2_e=of_g_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==3)))]

Approach1_o_aep=aep_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==0)))]
Approach2_o_aep=aep_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==1)))]
Approach1_e_aep=aep_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==2)))]
Approach2_e_aep=aep_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==3)))]

Approach1_o_cost_global=Glob_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==0)))]
Approach2_o_cost_global=Glob_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==1)))]
Approach1_e_cost_global=Glob_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==2)))]
Approach2_e_cost_global=Glob_array[np.argwhere(np.logical_and(Time_array>35,np.logical_and(Time_array<37,ide_array==3)))]

plt.figure()
plt.boxplot(np.concatenate((Approach1_o,Approach2_o),1),0,positions=x)
plt.xticks(x,labels,fontsize=31) 
plt.tick_params(axis="y", labelsize=31)
plt.ylabel('IRR [%]',fontsize=31)
plt.yticks(np.round(np.linspace(min(Approach2_o)-0.11, max(Approach2_o)+0.1, 10),3))
plt.ylim((min(Approach2_o)-0.11,max(Approach2_o)+0.1))
plt.show()

plt.figure()
plt.boxplot(np.concatenate((Approach1_e,Approach2_e),1),0,positions=x)
plt.xticks(x,labels,fontsize=31) 
plt.tick_params(axis="y", labelsize=31)
plt.ylabel('IRR [%]',fontsize=31)
plt.yticks(np.round(np.linspace(min(Approach2_e)-0.11, max(Approach2_e)+0.1, 10),3))
plt.ylim((min(Approach2_e)-0.11,max(Approach2_e)+0.1))
plt.show()

x2 = 2*np.arange(len(labels))
fig,ax = plt.subplots()
bplot1=ax.boxplot(np.concatenate((Approach1_o_aep,Approach2_o_aep),1),0,positions=x2-0.5,patch_artist=True)
colors = ['blue','blue']
for patch, color in zip(bplot1['boxes'], colors):
    patch.set_facecolor(color)
ax2=ax.twinx()
bplot2=ax2.boxplot(np.concatenate((Approach1_o_cost_global/1e6,Approach2_o_cost_global/1e6),1),0,positions=x2+0.5,patch_artist=True)
plt.xticks(x2,labels,fontsize=45) 
colors = ['red','red']
for patch, color in zip(bplot2['boxes'], colors):
    patch.set_facecolor(color)
ax.tick_params(axis="y", labelsize=45,colors='blue')
ax.set_yticks(list(np.round(np.linspace(min(Approach2_o_aep)[0]-5, max(Approach1_o_aep)[0]+5, 10),0)), minor=False)
ax.set_ylim((min(Approach2_o_aep)[0]-5,max(Approach1_o_aep)[0]+4.9))
ax.set_ylabel('AEP [GWh]',fontsize=45)
ax.yaxis.label.set_color('blue')  

ax2.tick_params(axis="y", labelsize=45,colors='red')
ax2.set_ylabel('Cables cost glob. opt. [Mill. euros]',fontsize=45)
ax2.yaxis.label.set_color('red')  
ax2.set_yticks(list(np.round(np.linspace(min(Approach2_o_cost_global/1e6)[0]-5, max(Approach1_o_cost_global/1e6)[0]+5, 10),0)), minor=False)
ax2.set_ylim((min(Approach2_o_cost_global/1e6)[0]-5.6,max(Approach1_o_cost_global/1e6)[0]+5.5))

ax.tick_params(axis="x", labelsize=45)
plt.show()

x2 = 2*np.arange(len(labels))
fig,ax = plt.subplots()
bplot1=ax.boxplot(np.concatenate((Approach1_e_aep,Approach2_e_aep),1),0,positions=x2-0.5,patch_artist=True)
colors = ['blue','blue']
for patch, color in zip(bplot1['boxes'], colors):
    patch.set_facecolor(color)
ax2=ax.twinx()
bplot2=ax2.boxplot(np.concatenate((Approach1_e_cost_global/1e6,Approach2_e_cost_global/1e6),1),0,positions=x2+0.5,patch_artist=True)
plt.xticks(x2,labels,fontsize=45) 
colors = ['red','red']
for patch, color in zip(bplot2['boxes'], colors):
    patch.set_facecolor(color)
ax.tick_params(axis="y", labelsize=45,colors='blue')
ax.set_yticks(list(np.round(np.linspace(min(Approach2_e_aep)[0]-5, max(Approach1_e_aep)[0]+5, 10),0)), minor=False)
ax.set_ylim((min(Approach2_e_aep)[0]-5,max(Approach1_e_aep)[0]+4.9))
ax.set_ylabel('AEP [GWh]',fontsize=45)
ax.yaxis.label.set_color('blue')  
ax2.tick_params(axis="y", labelsize=45,colors='red')
ax2.set_ylabel('Cables cost glob. opt. [Mill. euros]',fontsize=45)
ax2.yaxis.label.set_color('red')  

ax2.set_yticks(list(np.round(np.linspace(min(Approach2_e_cost_global/1e6)[0]-5, max(Approach1_e_cost_global/1e6)[0]+5, 10),0)), minor=False)
ax2.set_ylim((min(Approach2_e_cost_global/1e6)[0]-5.4,max(Approach1_e_cost_global/1e6)[0]+5.1))

ax.tick_params(axis="x", labelsize=45)
plt.show()







x2 = 2*np.arange(len(labels))
fig,ax = plt.subplots()
bplot1=ax.boxplot(np.concatenate((Approach1_o_aep,Approach2_o_aep),1),0,positions=x2-0.5,patch_artist=True)
colors = ['blue','blue']
for patch, color in zip(bplot1['boxes'], colors):
    patch.set_facecolor(color)
ax2=ax.twinx()
bplot2=ax2.boxplot(np.concatenate((Approach1_o_cost_global/1e6,Approach2_o_cost_global/1e6),1),0,positions=x2+0.5,patch_artist=True)
plt.xticks(x2,labels,fontsize=45) 
colors = ['red','red']
for patch, color in zip(bplot2['boxes'], colors):
    patch.set_facecolor(color)
ax.tick_params(axis="y", labelsize=45,colors='blue')
ax.set_yticks(list(np.round(np.linspace(min(Approach2_e_aep)[0]-5, max(Approach1_e_aep)[0]+5, 10),0)), minor=False)
ax.set_ylim((min(Approach2_e_aep)[0]-5,max(Approach1_e_aep)[0]+4.9))
ax.set_ylabel('AEP [GWh]',fontsize=45)
ax.yaxis.label.set_color('blue')  

ax2.tick_params(axis="y", labelsize=45,colors='red')
ax2.set_ylabel('Cables cost glob. opt. [Mill. euros]',fontsize=45)
ax2.yaxis.label.set_color('red')  
ax2.set_yticks(list(np.round(np.linspace(min(Approach2_o_cost_global/1e6)[0]-5, max(Approach1_o_cost_global/1e6)[0]+5, 10),0)), minor=False)
ax2.set_ylim((min(Approach2_o_cost_global/1e6)[0]-5.6,max(Approach1_o_cost_global/1e6)[0]+5.5))

ax.tick_params(axis="x", labelsize=45)
plt.show()

x2 = 2*np.arange(len(labels))
fig,ax = plt.subplots()
bplot1=ax.boxplot(np.concatenate((Approach1_e_aep,Approach2_e_aep),1),0,positions=x2-0.5,patch_artist=True)
colors = ['blue','blue']
for patch, color in zip(bplot1['boxes'], colors):
    patch.set_facecolor(color)
ax2=ax.twinx()
bplot2=ax2.boxplot(np.concatenate((Approach1_e_cost_global/1e6,Approach2_e_cost_global/1e6),1),0,positions=x2+0.5,patch_artist=True)
plt.xticks(x2,labels,fontsize=45) 
colors = ['red','red']
for patch, color in zip(bplot2['boxes'], colors):
    patch.set_facecolor(color)
ax.tick_params(axis="y", labelsize=45,colors='blue')
ax.set_yticks(list(np.round(np.linspace(min(Approach2_e_aep)[0]-5, max(Approach1_e_aep)[0]+5, 10),0)), minor=False)
ax.set_ylim((min(Approach2_e_aep)[0]-5,max(Approach1_e_aep)[0]+4.9))
ax.set_ylabel('AEP [GWh]',fontsize=45)
ax.yaxis.label.set_color('blue')  
ax2.tick_params(axis="y", labelsize=45,colors='red')
ax2.set_ylabel('Cables cost glob. opt. [Mill. euros]',fontsize=45)
ax2.yaxis.label.set_color('red')  

ax2.set_yticks(list(np.round(np.linspace(min(Approach2_e_cost_global/1e6)[0]-5, max(Approach1_e_cost_global/1e6)[0]+5, 10),0)), minor=False)
ax2.set_ylim((min(Approach2_e_cost_global/1e6)[0]-5.4,max(Approach1_e_cost_global/1e6)[0]+5.1))

ax.tick_params(axis="x", labelsize=45)
plt.show()










plt.figure()
plt.boxplot(np.concatenate((Approach1_o_aep,Approach2_o_aep),1),0,positions=x)
plt.xticks(x,labels,fontsize=31) 
plt.tick_params(axis="y", labelsize=31)
plt.ylabel('AEP [GWh]',fontsize=31)
plt.yticks(np.round(np.linspace(min(Approach2_o_aep)-1, max(Approach1_o_aep)+1, 10),3))
plt.ylim((min(Approach2_o_aep)-1.2,max(Approach1_o_aep)+1))
plt.show()

plt.figure()
plt.boxplot(np.concatenate((Approach1_o_cost_global/1e6,Approach2_o_cost_global/1e6),1),0,positions=x)
plt.xticks(x,labels,fontsize=31) 
plt.tick_params(axis="y", labelsize=31)
plt.ylabel('Cost from global optimizer [Millions of euros]',fontsize=31)
plt.yticks(np.round(np.linspace(min(Approach2_o_cost_global/1e6)-5, max(Approach1_o_cost_global/1e6)+6, 10),3))
plt.ylim((min(Approach2_o_cost_global/1e6)-5,max(Approach1_o_cost_global/1e6)+6.1))
plt.show()

plt.figure()
plt.boxplot(np.concatenate((Approach1_e_aep,Approach2_e_aep),1),0,positions=x)
plt.xticks(x,labels,fontsize=31) 
plt.tick_params(axis="y", labelsize=31)
plt.ylabel('AEP [GWh]',fontsize=31)
plt.yticks(np.round(np.linspace(min(Approach2_e_aep)-1, max(Approach1_e_aep)+1, 10),3))
plt.ylim((min(Approach2_e_aep)-1.2,max(Approach1_e_aep)+1))
plt.show()

plt.figure()
plt.boxplot(np.concatenate((Approach1_e_cost_global/1e6,Approach2_e_cost_global/1e6),1),0,positions=x)
plt.xticks(x,labels,fontsize=31) 
plt.tick_params(axis="y", labelsize=31)
plt.ylabel('Cost from global optimizer [Millions of euros]',fontsize=31)
plt.yticks(np.round(np.linspace(min(Approach2_e_cost_global/1e6)-5, max(Approach1_e_cost_global/1e6)+6, 10),3))
plt.ylim((min(Approach2_e_cost_global/1e6)-5,max(Approach1_e_cost_global/1e6)+6.1))
plt.show()

plt.figure()
plt.plot(Glob_array[ide_array==2]/1e6,Heu_array[ide_array==2]/1e6,'o',markersize=15,color='b',marker='^')
m, b = np.polyfit(Glob_array[ide_array==2]/1e6,Heu_array[ide_array==2]/1e6, 1)
plt.plot(Glob_array[ide_array==2]/1e6, (m*Glob_array[ide_array==2]/1e6 + b),color="orange")
plt.xlabel('Cost from global optimizer [Millions of euros]',fontsize=40)
plt.ylabel('Cost from heuristic [Millions of euros]',fontsize=40)
plt.yticks(np.round(np.linspace(min(Heu_array[ide_array==2]/1e6)-5, max(Heu_array[ide_array==2]/1e6)+5, 10),0))
plt.tick_params(axis="y", labelsize=40)
plt.tick_params(axis="x", labelsize=40)
plt.ylim((min(Heu_array[ide_array==2]/1e6)-5.2,max(Heu_array[ide_array==2]/1e6)+5))
#plt.title('Correlation global optimizer and heuristic for 50th percentile of cost difference')
plt.text(1.225e2,2.1e2,'r='+str(round(np.corrcoef(Glob_array[ide_array==2],Heu_array[ide_array==2])[0,1],2)),fontsize=35)

plt.figure()
plt.plot(Glob_array[ide_array==1]/1e6,Heu_array[ide_array==1]/1e6,'o',markersize=15,color='r')
m, b = np.polyfit(Glob_array[ide_array==1]/1e6,Heu_array[ide_array==1]/1e6, 1)
plt.plot(Glob_array[ide_array==1]/1e6, (m*Glob_array[ide_array==1]/1e6 + b),color="orange")
plt.xlabel('Cost from global optimizer [Millions of euros]',fontsize=40)
plt.ylabel('Cost from heuristic [Millions of euros]',fontsize=40)
plt.yticks(np.round(np.linspace(min(Heu_array[ide_array==1]/1e6)-5, max(Heu_array[ide_array==1]/1e6)+5, 10),0))
plt.tick_params(axis="y", labelsize=40)
plt.tick_params(axis="x", labelsize=40)
plt.ylim((min(Heu_array[ide_array==1]/1e6)-4.8,max(Heu_array[ide_array==1]/1e6)+5.3))
#plt.title('Correlation global optimizer and heuristic for 50th percentile of cost difference')
plt.text(0.825e2,1.25e2,'r='+str(round(np.corrcoef(Glob_array[ide_array==1],Heu_array[ide_array==1])[0,1],2)),fontsize=35)

plt.figure()
plt.plot(Length_av_array[np.argwhere(ide_array==0)],Perc_array[np.argwhere(ide_array==0)],'o',markersize=15,color='b',label='OSS at approximately centroid, Layouts from Approach 1')
plt.plot(Length_av_array[np.argwhere(ide_array==1)],Perc_array[np.argwhere(ide_array==1)],'o',markersize=15,color='r',label='OSS at approximately centroid, Layouts from Approach 2')
plt.plot(Length_av_array[np.argwhere(ide_array==2)],Perc_array[np.argwhere(ide_array==2)],'^',markersize=15,color='b',label='OSS at external location, Layouts from Approach 1')
plt.plot(Length_av_array[np.argwhere(ide_array==3)],Perc_array[np.argwhere(ide_array==3)],'^',markersize=15,color='r',label='OSS at external location, Layouts from Approach 2')
m, b = np.polyfit(Length_av_array,Perc_array, 1)
plt.plot(Length_av_array, (m*Length_av_array + b),color="orange")
plt.xlabel('Average length between pairs of WTs [D]',fontsize=40)
plt.ylabel('Absolute relative cost difference between'"\n"'heuristic and global optimizer [%]',fontsize=40)
plt.yticks(np.round(np.linspace(min(Perc_array)-1, max(Perc_array)+1, 10),0))
plt.tick_params(axis="y", labelsize=40)
plt.tick_params(axis="x", labelsize=40)
plt.ylim((min(Perc_array)-0.97,max(Perc_array)+1))
plt.legend(loc='upper left',fontsize=28)
plt.text(38.2,42.3,'r='+str(round(np.corrcoef(Length_av_array,Perc_array)[0,1],2)),fontsize=35)
plt.show() 

plt.figure()
plt.plot(Length_sd_array[np.argwhere(ide_array==0)],Perc_array[np.argwhere(ide_array==0)],'o',markersize=15,color='b',label='OSS at approximately centroid, Layouts from Approach 1')
plt.plot(Length_sd_array[np.argwhere(ide_array==1)],Perc_array[np.argwhere(ide_array==1)],'o',markersize=15,color='r',label='OSS at approximately centroid, Layouts from Approach 2')
plt.plot(Length_sd_array[np.argwhere(ide_array==2)],Perc_array[np.argwhere(ide_array==2)],'^',markersize=15,color='b',label='OSS at external location, Layouts from Approach 1')
plt.plot(Length_sd_array[np.argwhere(ide_array==3)],Perc_array[np.argwhere(ide_array==3)],'^',markersize=15,color='r',label='OSS at external location, Layouts from Approach 2')
m, b = np.polyfit(Length_sd_array,Perc_array, 1)
plt.plot(Length_sd_array, (m*Length_sd_array + b),color="orange")
plt.xlabel('Standard deviation length between pairs of WTs [D]',fontsize=40)
plt.ylabel('Absolute relative cost difference between'"\n"'heuristic and global optimizer [%]',fontsize=40)
plt.yticks(np.round(np.linspace(min(Perc_array)-1, max(Perc_array)+1, 10),0))
plt.tick_params(axis="y", labelsize=40)
plt.tick_params(axis="x", labelsize=40)
plt.ylim((min(Perc_array)-0.97,max(Perc_array)+1))
plt.legend(loc='upper left',fontsize=28)
plt.text(19.1,42.5,'r='+str(round(np.corrcoef(Length_sd_array,Perc_array)[0,1],2)),fontsize=35)
plt.show() 

plt.figure()
plt.plot(Glob_array[np.argwhere(ide_array==0)]/1e6,aep_array[np.argwhere(ide_array==0)],'o',markersize=15,color='b',label='OSS at approximately centroid, Layouts from Approach 1')
plt.plot(Glob_array[np.argwhere(ide_array==1)]/1e6,aep_array[np.argwhere(ide_array==1)],'o',markersize=15,color='r',label='OSS at approximately centroid, Layouts from Approach 2')
#plt.plot(Glob_array[np.argwhere(ide_array==2)],aep_array[np.argwhere(ide_array==2)],'^',markersize=15,color='b',label='OSS at external location, Layouts from Approach 1')
#plt.plot(Glob_array[np.argwhere(ide_array==3)],aep_array[np.argwhere(ide_array==3)],'^',markersize=15,color='r',label='OSS at external location, Layouts from Approach 2')
m, b = np.polyfit(Glob_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))][:,0]/1e6,aep_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))][:,0], 1)
plt.plot(Glob_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))]/1e6, (m*Glob_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))]/1e6 + b),color="orange")
plt.xlabel('Cost from global optimizer [Millions of euros]',fontsize=40)
plt.ylabel('AEP [GWh]',fontsize=40)
plt.yticks(np.round(np.linspace(min(aep_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))])-1, max(aep_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))])+1, 10),0))
plt.tick_params(axis="y", labelsize=40)
plt.tick_params(axis="x", labelsize=40)
plt.ylim((min(aep_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))])-0.97,max(aep_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))])+1))
plt.legend(loc='upper left',fontsize=28)
plt.text(85,3788,'r='+str(round(np.corrcoef(Glob_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))][:,0]/1e6,aep_array[np.argwhere(np.logical_or(ide_array==0,ide_array==1))][:,0])[0,1],2)),fontsize=35)
plt.show() 


#plt.title('Correlation global optimizer and heuristic for 50th percentile of cost difference')
#plt.text(0.825e2,1.25e2,'r='+str(round(np.corrcoef(Glob_array[ide_array==1],Heu_array[ide_array==1])[0,1],2)),fontsize=31)

