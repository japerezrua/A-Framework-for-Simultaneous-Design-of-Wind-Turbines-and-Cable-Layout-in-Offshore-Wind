# -*- coding: utf-8 -*-
"""
Created on Mon Jan 25 07:57:12 2021

@author: juru
"""

import openpyxl
import numpy as np
from global_collection import global_optimizer
from plotting_global import plotting
from collection_system import collection_system
from cables_cost_function import cost_function_array
from topfarm.cost_models.economic_models.dtu_wind_cm_main import economic_evaluation as EE_DTU
from borssele import get_site
from borssele import IEA10MW

row=10

site = get_site()
n_wt = len(site.initial_position)
windTurbines = IEA10MW()
wt_power=10000 #Wind turbine Voltage in kilowatt
Drotor_vector = [windTurbines.diameter()] * n_wt 
power_rated_vector = wt_power* np.ones(n_wt)
hub_height_vector = [windTurbines.hub_height()] * n_wt 
rated_rpm_array = 8.68 * np.ones([n_wt])
n_wd = 16
wind_directions = np.linspace(0., 360., n_wd, endpoint=False)

from py_wake.wake_models.gaussian import IEA37SimpleBastankhahGaussian
from py_wake.aep_calculator import AEPCalculator

## We use the Gaussian wake model
wake_model = IEA37SimpleBastankhahGaussian(site, windTurbines)
AEPCalc = AEPCalculator(wake_model)

C=10
Cables=np.array([[500,4],[800,6],[1000,8]]) #Set of cables available for the full problem
option=3 #Choose heuristic for the cable layout. 1 = Prim. 2 = Kruskal. 3 = Esau-Williams
max_it=1000000 #Maximum number of iterations of the heuristics
Inters_const=False #Activate non-cables crossing
voltage=33 #Voltage level in kV of the collection system (33 kV, 66 kV)
power_rated_vector = wt_power* np.ones(n_wt)

oss=np.array([492222.525,5723736.425])

distance_from_shore = 10.0 # [km]
energy_price = 0.2 / 7.4 # [DKK/kWh] / [DKK/EUR] -> [EUR/kWh]
project_duration = 20 # [years]
water_depth_array = 20 * np.ones([n_wt])
Power_rated_array = np.array(power_rated_vector)/1.0E3 # [MW]

ee_dtu = EE_DTU(distance_from_shore, energy_price, project_duration)


def irr_dtu(aep, electrical_connection_cost, **kwargs):
    ee_dtu.calculate_irr(
                    rated_rpm_array, 
                    Drotor_vector, 
                    Power_rated_array,
                    hub_height_vector, 
                    water_depth_array, 
                    aep, 
                    electrical_connection_cost)
    return ee_dtu.IRR

wb = openpyxl.load_workbook('Results_compilation - OSS_o.xlsx')
sheet = wb['Approach 1 (WithoutCables)']
a=sheet.cell(row=row, column=1).value
a+='.xlsx'
print(a)

wb2 = openpyxl.load_workbook(a)
sheet2 = wb2['WTs positions']

X=np.zeros(n_wt)
Y=np.zeros(n_wt)

for i in range(3,77):
        X[i-3]=sheet2.cell(row=i, column=1).value
        Y[i-3]=sheet2.cell(row=i, column=2).value
        
X2=np.insert(X,0,oss[0])
Y2=np.insert(Y,0,oss[1])

Cables=cost_function_array(wt_power,voltage,Cables) 
CablesForGlobal=Cables 
ULForCables=max(CablesForGlobal[:,1]).item()


Tree,Cost_collection,gap_outputit,time_formulating,time_solving,solutions=global_optimizer(WTc=n_wt,X=X2,Y=Y2,
Cables=CablesForGlobal,C=C)
plotting(X2,Y2,CablesForGlobal,Tree)

T2,elnet_layout2 = collection_system(X2,Y2,option,ULForCables,Inters_const,max_it,CablesForGlobal,True)

aep_final=AEPCalc.calculate_AEP(x_i=X, y_i=Y, wd=wind_directions).sum(-1).sum(-1)*10**6

obj_heu=ee_dtu.calculate_irr(rated_rpm_array, Drotor_vector, Power_rated_array,hub_height_vector, water_depth_array, 
                    aep_final, elnet_layout2)
obj_global=ee_dtu.calculate_irr(rated_rpm_array, Drotor_vector, Power_rated_array,hub_height_vector, water_depth_array, 
                    aep_final, Cost_collection)

wb3 = openpyxl.load_workbook('Results_compilation - OSS_o.xlsx')
sheet3=wb3["Approach 1 (WithoutCables)"]
sheet3.cell(row=row, column=3).value=obj_heu
sheet3.cell(row=row, column=4).value=obj_global
sheet3.cell(row=row, column=5).value=sum(aep_final)/(10**6)
sheet3.cell(row=row, column=6).value=Cost_collection
sheet3.cell(row=row, column=7).value=elnet_layout2
wb3.save('Results_compilation - OSS_o.xlsx')

print("Objective cost with heuristic using full set of cables available:",obj_heu)

print("Objective cost with global optimizer using full set of cables available:",obj_global)

print("AEP in GWh:",sum(aep_final)/(10**6))
print("Cost of cable layout in MEuro of the global optimizer:",Cost_collection/(10**6))
print("Cost of cable layout in MEuro of the heuristic:",elnet_layout2/(10**6))