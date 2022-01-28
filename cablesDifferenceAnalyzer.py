# -*- coding: utf-8 -*-
"""
Created on Tue Mar  9 12:58:16 2021

@author: juru
"""

import openpyxl
import numpy as np
import statistics
import matplotlib.pylab as plt

#%% Inputs
n_wt=74

wb1 = openpyxl.load_workbook('Results_compilation.xlsx',data_only=True)
wb2 = openpyxl.load_workbook('Results_compilation - OSS_o.xlsx',data_only=True)

sheet11 = wb1['Approach 1 (WithoutCables)']
sheet12 = wb1['Approach 2 (WithCables)']

sheet21 = wb2['Approach 1 (WithoutCables)']
sheet22 = wb2['Approach 2 (WithCables)']

sheets=[sheet11,sheet12,sheet21,sheet22]



#%% Process
stats_mean,stats_dev,diff_perc=np.empty(0),np.empty(0),np.empty(0)
k_array,cost_global,cost_heu=np.empty(0),np.empty(0),np.empty(0)
name,of_d_array,of_g_array,time,aep_array=[],np.empty(0),np.empty(0),np.empty(0),np.empty(0)
cables_array_g,cables_array_h,k_array_2,name_2=np.empty((0,3)),np.empty((0,3)),np.empty(0),[]
for k in range(len(sheets)):
    sheet=sheets[k]
    for i in range(2,1000,1):
        a=sheet.cell(row=i, column=1).value
        if a is None:
           break
        diff=sheet.cell(row=i, column=11).value
        glob=sheet.cell(row=i, column=6).value
        heu=sheet.cell(row=i, column=7).value
        aep=sheet.cell(row=i, column=5).value
        if k==0:
           glob,heu=1000*glob,1000*heu
        of_d=sheet.cell(row=i, column=2).value
        of_g=sheet.cell(row=i, column=4).value
        time1=sheet.cell(row=i, column=8).value
        X=np.zeros(n_wt)
        Y=np.zeros(n_wt)
        a+='.xlsx'
        wb3 = openpyxl.load_workbook(a,data_only=True)
        sheet3 = wb3['WTs positions']
        sheet4 = wb3['Global optimizer cable design']
        sheet5 = wb3['Heur. optimizer cable design']         
        for j in range(3,n_wt+3):
            X[j-3]=sheet3.cell(row=j, column=1).value
            Y[j-3]=sheet3.cell(row=j, column=2).value
        dist_to_wt=np.empty(0)
        for j in range(n_wt):
            X_c,Y_c=np.copy(X[j+1:n_wt]),np.copy(Y[j+1:n_wt])
            dist_to_wt_2=np.sqrt((X_c-X[j])**2+(Y_c-Y[j])**2)
            dist_to_wt=np.concatenate((dist_to_wt,dist_to_wt_2))
        stats_mean=np.concatenate((stats_mean,np.array(statistics.mean(dist_to_wt)).reshape(1)))
        stats_dev=np.concatenate((stats_dev,np.array(statistics.stdev(dist_to_wt)).reshape(1)))
        diff_perc=np.concatenate((diff_perc,np.array(diff).reshape(1)))
        k_array=np.concatenate((k_array,np.array(k).reshape(1)))
        cost_global=np.concatenate((cost_global,np.array(glob).reshape(1)))
        cost_heu=np.concatenate((cost_heu,np.array(heu).reshape(1)))
        aep_array=np.concatenate((aep_array,np.array(aep).reshape(1)))
        of_d_array=np.concatenate((of_d_array,np.array(of_d).reshape(1)))
        of_g_array=np.concatenate((of_g_array,np.array(of_g).reshape(1)))
        time=np.concatenate((time,np.array(time1).reshape(1)))        
        name+=[a]
        if (k==0 or k==1):
            cables_array_t_g,cables_array_t_h=np.zeros((1,3)),np.zeros((1,3))
            for j in range(2,1000,1):
                aa1=sheet4.cell(row=j, column=4).value
                aa2=sheet5.cell(row=j, column=4).value
                if aa1 is None:
                   break                
                if (aa1==0):
                    cables_array_t_g[0,0]+=sheet4.cell(row=j, column=3).value
                if (aa1==1):
                    cables_array_t_g[0,1]+=sheet4.cell(row=j, column=3).value
                if (aa1==2):
                    cables_array_t_g[0,2]+=sheet4.cell(row=j, column=3).value
                if (aa2==0):
                    cables_array_t_h[0,0]+=sheet5.cell(row=j, column=3).value
                if (aa2==1):
                    cables_array_t_h[0,1]+=sheet5.cell(row=j, column=3).value
                if (aa2==2):
                    cables_array_t_h[0,2]+=sheet5.cell(row=j, column=3).value
            cables_array_g=np.concatenate((cables_array_g,cables_array_t_g),axis=0)
            cables_array_h=np.concatenate((cables_array_h,cables_array_t_h),axis=0)                           
            k_array_2=np.concatenate((k_array_2,np.array(k).reshape(1)))
            name_2+=[a]              
args=np.argsort(stats_mean[:])
stats_mean=stats_mean[args]
diff_perc=100*diff_perc[args]
stats_dev=stats_dev[args]
k_array=k_array[args]
cost_global=cost_global[args]
cost_heu=cost_heu[args]
name=np.array(name)
name=name[args]
of_d_array=of_d_array[args]
aep_array=aep_array[args]
of_g_array=of_g_array[args]
time=time[args]
#%% Output saving
import xlsxwriter 
from datetime import datetime
now = datetime.now()
dt_string = now.strftime("%d-%m-%Y_%H-%M-%S")
workbook = xlsxwriter.Workbook(dt_string+'_Processing_'+'.xlsx')
worksheet0 = workbook.add_worksheet("Outputs")
worksheet0.write(0,0,"File name")
worksheet0.write(0,1,"Time [h]")    
worksheet0.write(0,2,"Length average [m]")
worksheet0.write(0,3,"Length standard deviation [m]")
worksheet0.write(0,4,"Percentual difference [%]")
worksheet0.write(0,5,"Cost heuristic [Euro]")
worksheet0.write(0,6,"Cost global [Euro]")
worksheet0.write(0,7,"AEP [GWh]")
worksheet0.write(0,8,"OF D [%]")
worksheet0.write(0,9,"OF G [%]")
worksheet0.write(0,10,"ID")
for i in range(len(stats_mean)):
    worksheet0.write(i+1,0,name[i])
    worksheet0.write(i+1,1,time[i])
    worksheet0.write(i+1,2,stats_mean[i])
    worksheet0.write(i+1,3,stats_dev[i])
    worksheet0.write(i+1,4,diff_perc[i])
    worksheet0.write(i+1,5,cost_heu[i])
    worksheet0.write(i+1,6,cost_global[i])
    worksheet0.write(i+1,7,aep_array[i])
    worksheet0.write(i+1,8,of_d_array[i])
    worksheet0.write(i+1,9,of_g_array[i])
    worksheet0.write(i+1,10,k_array[i])
worksheet1 = workbook.add_worksheet("Cables outputs")
worksheet1.write(0,0,"File name")
worksheet1.write(0,1,"ID")
worksheet1.write(0,2,"Global cable 500mm2")
worksheet1.write(0,3,"Global cable 800mm2")    
worksheet1.write(0,4,"Global cable 1000mm2")
worksheet1.write(0,5,"Heuristic cable 500mm2")
worksheet1.write(0,6,"Heuristic cable 800mm2")    
worksheet1.write(0,7,"Heuristic cable 1000mm2")       
for i in range(len(k_array_2)):
    worksheet1.write(i+1,0,name_2[i])
    worksheet1.write(i+1,1,k_array_2[i])
    worksheet1.write(i+1,2,cables_array_g[i,0])
    worksheet1.write(i+1,3,cables_array_g[i,1])       
    worksheet1.write(i+1,4,cables_array_g[i,2])       
    worksheet1.write(i+1,5,cables_array_h[i,0])
    worksheet1.write(i+1,6,cables_array_h[i,1])       
    worksheet1.write(i+1,7,cables_array_h[i,2])        
workbook.close()   